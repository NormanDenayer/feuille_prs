import datetime
import calendar

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from argparse import ArgumentParser
from copy import copy

MONTHS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre'] 
DAYS = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']


def get_right_cell_coordinate(x, n=1):
    c = column_index_from_string(x)
    return get_column_letter(c+n)


def main():
    parser = ArgumentParser()
    parser.add_argument('-s', '--source', default='rsrc/2018 November.xlsx')
    parser.add_argument('-y', '--year', type=int, required=True)
    args = parser.parse_args()

    wb = load_workbook(filename=args.source)
    base_ = wb.active
    we_fill = copy(base_['F5'].fill)
    nl_fill = copy(base_['E5'].fill)

    for i, month in enumerate(MONTHS, start=1):
        ws = wb.copy_worksheet(base_)
        ws.title = f'{month} {args.year}'
        _, last_day_of_the_month = calendar.monthrange(args.year, i)
        week_totals = []

        for nb_day in range(1, last_day_of_the_month+1):
            day = datetime.datetime(args.year, i, nb_day)
            
            col = get_right_cell_coordinate("C", nb_day)
            ws[f'{col}4'] = DAYS[day.weekday()]
            ws[f'{col}5'] = str(nb_day)
            
            for j in range(4, 9):
                ws[f'{col}{j}'].fill = we_fill if day.weekday() in (5, 6) else nl_fill
                if j > 5:
                    ws[f'{col}{j}'] = None
            if day.weekday() == 4 or (nb_day == last_day_of_the_month and day.weekday() < 5):
                if day.weekday() == 4:
                    first_day_of_the_week = max(nb_day-4, 1)
                else:
                    first_day_of_the_week = nb_day - day.weekday()
                ws[f'{col}7'] = f'=SUM({get_right_cell_coordinate("C", first_day_of_the_week)}6:{col}6)'
                week_totals.append(f'{col}7')
        ws['S2'] = f'{month} {args.year}'
        ws['AI7'] = f'=SUM({",".join(week_totals)})'

    wb.remove(base_)
    wb.save(f'feuille_{args.year}.xlsx')


if __name__ == '__main__':
    main()
